"""
Regenerate codemaps/*.csv from the official MoSPI documentation in docs/.

The code lists are published in two places and neither is machine-friendly:
  docs/CODEs for Blocks of Sch - CMS-Education.xlsx   — value labels, blocks 1–5
  docs/Data_Layout_CMSE_2025.xlsx  ("State code" sheet) — state code -> name

This script flattens both into CSVs that the transform and any downstream
analyst can read. The authoritative dicts used by the pipeline live in
sources.py; this script emits the CSV mirror and the full raw label dump.

Usage:
  python3 scripts/build_codemaps.py
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl

import sources as S

CODES_XLSX = S.DOCS / "CODEs for Blocks of Sch - CMS-Education.xlsx"
LAYOUT_XLSX = S.DOCS / "Data_Layout_CMSE_2025.xlsx"


def _write(name: str, header: list[str], rows: list[list]) -> None:
    path = S.CODEMAPS / name
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)
    print(f"  wrote {path.name:38s} {len(rows):4d} rows")


def build_state_map() -> None:
    ws = openpyxl.load_workbook(LAYOUT_XLSX, data_only=True)["State code"]
    rows = [
        [str(code).strip().zfill(2), " ".join(str(name).split())]
        for code, name in ws.iter_rows(min_row=3, values_only=True)
        if code and name
    ]
    _write("cmse_state.csv", ["state_code", "state_name"], rows)


def build_value_labels() -> None:
    """Full dump of every published value label, keyed by schedule block + item."""
    wb = openpyxl.load_workbook(CODES_XLSX, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        for r in wb[sheet].iter_rows(min_row=3, values_only=True):
            block, item, var, code, desc = r[0], r[1], r[2], r[3], r[4]
            if not (var and code is not None and desc):
                continue
            rows.append([
                str(block), str(item), " ".join(str(var).split()),
                str(code).strip(), " ".join(str(desc).split()),
            ])
    _write("cmse_value_labels.csv", ["block", "item", "variable_label", "code", "code_description"], rows)


def build_pipeline_codemaps() -> None:
    """CSV mirror of the decode maps the transform actually applies."""
    simple = {
        "cmse_sector.csv": ("sector_code", "sector_name", S.SECTOR),
        "cmse_gender.csv": ("gender_code", "gender_name", S.GENDER),
        "cmse_social_group.csv": ("social_group_code", "social_group_name", S.SOCIAL_GROUP),
        "cmse_religion.csv": ("religion_code", "religion_name", S.RELIGION),
        "cmse_relation_to_head.csv": ("relation_code", "relation_name", S.RELATION_TO_HEAD),
        "cmse_enrolment_level.csv": ("enrolment_level_code", "enrolment_level_name", S.ENROLMENT_LEVEL),
        "cmse_school_type.csv": ("school_type_code", "school_type_name", S.SCHOOL_TYPE),
        "cmse_funding_source.csv": ("funding_source_code", "funding_source_name", S.FUNDING_SOURCE),
        "cmse_residence_type.csv": ("residence_type_code", "residence_type_name", S.RESIDENCE_TYPE),
    }
    for fname, (ccol, ncol, mapping) in simple.items():
        rows = [[str(k).zfill(2) if fname.endswith(("level.csv", "source.csv")) else str(k), v]
                for k, v in mapping.items()]
        _write(fname, [ccol, ncol], rows)

    # Household type is sector-dependent — one file, sector as a key column.
    rows = [["1", str(k), v] for k, v in S.HOUSEHOLD_TYPE_RURAL.items()]
    rows += [["2", str(k), v] for k, v in S.HOUSEHOLD_TYPE_URBAN.items()]
    _write("cmse_household_type.csv", ["sector_code", "household_type_code", "household_type_name"], rows)

    rows = [["school", str(k), v] for k, v in S.NONREPORTING_SCHOOL.items()]
    rows += [["coaching", str(k), v] for k, v in S.NONREPORTING_COACHING.items()]
    rows += [["boarding", str(k), v] for k, v in S.NONREPORTING_BOARDING.items()]
    _write("cmse_nonreporting_code.csv", ["item", "code", "code_description"], rows)


def main() -> None:
    S.CODEMAPS.mkdir(exist_ok=True)
    print(f"CMS-E codemaps → {S.CODEMAPS}")
    build_state_map()
    build_value_labels()
    build_pipeline_codemaps()
    print("done.")


if __name__ == "__main__":
    main()
