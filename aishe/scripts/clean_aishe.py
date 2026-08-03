"""
Parse the AISHE Final Reports in raw/ into the single denormalized higher-ed
fact (clean/higher_ed.parquet → BQ aishe_fact_higher_ed_students).

This is the one entry point for the fact. It reads BOTH published editions:

  Excel workbooks (2019-20 onward)   parsed here, via openpyxl
  PDF reports     (2012-13 … 2018-19) parsed by parse_report_pdf.py, which this
                                      script calls — those years have no Excel
                                      edition at all

The fact unifies several published cuts into one grain. Every row is tagged
with `cut` (which published slice it came from) and `metric` (enrolment vs
graduates). Dimensions a given cut doesn't break out carry the sentinel "All".
ALWAYS filter on `cut`; the cuts overlap, so SUMMing across them double-counts.

  cut='state_level'      Table 33      graduates by state x level
  cut='programme_social' Table 34a     graduates by programme x social category
  cut='ug_discipline'    Tables 12 + 35  UG enrolment (T12) + graduates (T35) by
                                        discipline (the multi-year trend)

metric='enrolment' exists only on the ug_discipline cut (from Table 12);
metric='graduates' exists on all three cuts.

Which (year, sheet) pairs are parsed — and therefore which cuts a year
contributes — is declared in sources.RAW_SHEETS (Excel) and sources.PDF_TABLES
(PDF), not here.

Grain: (cut, aishe_year, metric, level, state, discipline, programme,
        social_category, gender) -> value

**On partial builds.** load_bq.py loads this table with WRITE_TRUNCATE, so the
parquet is the whole table — writing a subset of the years silently DELETES the
rest from BigQuery. The Excel workbooks cannot currently be re-downloaded (see
sources.REPORT_URLS), so a from-scratch checkout can only rebuild the PDF years.
Rather than let that quietly truncate the table, a build that is missing any
registered Excel year refuses to write clean/higher_ed.parquet at all; with
--allow-missing-excel it writes clean/higher_ed.partial.parquet instead, a
filename upload_to_gcs.py and load_bq.py never read.

Usage:
  python3 scripts/clean_aishe.py
  python3 scripts/clean_aishe.py --allow-missing-excel   # PDF years only
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
import parse_report_pdf
from sources import (BASIS_ACTUAL, CLEAN, PDF_TABLES, RAW_SHEETS, REPORTS,
                     SENTINEL, TABLES, canonical_state)

COLUMNS = ["cut", "aishe_year", "metric", "basis", "level", "state",
           "discipline", "programme", "social_category", "gender", "value"]

LEVELS = [
    "Ph.D.", "M.Phil.", "Post Graduate", "Under Graduate",
    "PG Diploma", "Diploma", "Certificate", "Integrated",
]
# Table 33 prints a 9th group after the levels: the row's roll-up across all of
# them. It is read (and used as a check) but never emitted as a level.
GRAND_TOTAL = "Grand Total"
STATE_LEVEL_GROUPS = LEVELS + [GRAND_TOTAL]
GENDERS = ["Male", "Female", "Total"]
SOCIAL_CATEGORIES = [
    "All Categories", "Scheduled Caste", "Scheduled Tribe",
    "Other Backward Classes", "Persons with Disability", "Muslim",
    "Other Minority Communities", "EWS",
]

# Published anchor: UG graduates (gender='Total') for a year, which must also
# reconcile between the state_level and ug_discipline cuts. The cross-cut
# equality is a self-check that holds for any year; this dict is the external
# figure to check against where we have one. Add a year's anchor when adding
# the year.
UG_GRADUATES_ANCHOR: dict[str, int] = {"2021-22": 7_754_223}

# Regression guard — row counts these cuts produced for 2021-22 before the
# geometry detection below replaced hardcoded column offsets. A change here
# means the parse moved; investigate before shipping.
ROW_BASELINE: dict[tuple[str, str], int] = {
    ("state_level", "2021-22"): 864,
    ("programme_social", "2021-22"): 5448,
}

# Cells that legitimately mean "no value" in these tables.
NULLISH = {"", "-", "--", "na", "n.a.", "n/a", "nil", "*", ".", "…"}


def _wb(year: str):
    path = REPORTS[year]
    if not path.exists():
        raise SystemExit(
            f"missing raw workbook: {path}\n"
            f"Restore it from the GCS mirror:  python3 scripts/fetch.py --from-gcs\n"
            f"(`--year {year}` will NOT work — the upstream Excel URLs all 404, so "
            f"the bucket is the only source. See sources.REPORT_URLS.)\n"
            f"To build the PDF-only years without it: "
            f"python3 scripts/clean_aishe.py --allow-missing-excel"
        )
    return openpyxl.load_workbook(path, data_only=True)


def _sheet(wb, *names):
    want = {n.replace(" ", "").lower() for n in names}
    for s in wb.sheetnames:
        if s.replace(" ", "").lower() in want:
            return wb[s]
    raise SystemExit(
        f"no sheet matching {names} (have: {wb.sheetnames})\n"
        f"AISHE renumbers tables between editions — run "
        f"`python3 scripts/inspect_workbook.py` and fix the sheet name in "
        f"sources.RAW_SHEETS."
    )


def _num(value, ctx: str) -> int:
    """Coerce a value cell to int, failing loudly on anything unexpected.

    Blank and the conventional 'no value' markers become 0. A numeric string
    (possibly comma-grouped) is parsed. Anything else — most importantly a text
    label, which is what lands in a value column once the layout shifts — is an
    error rather than a silent 0.
    """
    if value is None:
        return 0
    if isinstance(value, bool):
        raise SystemExit(f"unexpected boolean in a value column ({ctx})")
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if s.lower() in NULLISH:
        return 0
    try:
        return int(float(s.replace(",", "")))
    except ValueError:
        raise SystemExit(
            f"non-numeric value {value!r} in a value column ({ctx}).\n"
            f"The column layout has probably shifted. Run "
            f"`python3 scripts/inspect_workbook.py` to see the real geometry."
        )


def _row(cut, year, metric, level, state, discipline, programme,
         social_category, gender, value, ctx="", basis=BASIS_ACTUAL):
    """Build one fact row.

    `basis` defaults to actual-response because every sheet the Excel parser
    reads is a "based on actual response" table; the estimated tables only enter
    via the PDF parser (Tables 14/15).
    """
    return {
        "cut": cut, "aishe_year": year, "metric": metric, "basis": basis,
        "level": level, "state": state, "discipline": discipline,
        "programme": programme, "social_category": social_category,
        "gender": gender, "value": _num(value, ctx),
    }


def _key(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


ROLLUP_LABELS = {"grand", "grandtotal", "allindia", "india", "total"}


def _label_with_merge(row, label_col: int) -> str:
    """The row's label, tolerating the merged cell AISHE uses on its roll-up row.

    On every data row the label sits in `label_col` and the serial number to its
    left. On the *roll-up* row ("All India" / "Grand Total") the sheet merges the
    serial and label cells, so openpyxl reports the label in the SERIAL column and
    `label_col` comes back empty — which made the reader skip the row, leaving
    both published-total checks with no anchor and silently degrading them to a
    warning. Falling back one column left recovers it.

    The fallback only accepts non-numeric text: on an ordinary row the cell to the
    left holds the serial, and returning that would invent a state called "35".
    """
    label = row[label_col] if len(row) > label_col else None
    if label is not None and str(label).strip():
        return str(label).strip()
    if label_col == 0:
        return ""
    left = row[label_col - 1] if len(row) > label_col - 1 else None
    if left is None or isinstance(left, (int, float)):
        return ""
    text = str(left).strip()
    return text if text and not text.isdigit() else ""


# ─── Cross-tab geometry (Tables 33 / 34a) ────────────────────────────────────
def _crosstab_geometry(ws, expected_groups: list[str]) -> tuple[int, int, int]:
    """Locate a <label> × <group × Male/Female/Total> cross-tab's geometry.

    Returns (label_col, first_value_col, first_data_row).

    Reads the sheet rather than assuming it: the gender header row gives the
    value block's start and width, so an inserted S.No. column shifts the whole
    thing harmlessly. Raises if the block width doesn't match
    `expected_groups`, or if the group label row disagrees with it — the two
    ways an edition can invalidate the positional read without changing the
    sheet name.
    """
    genders = {"male", "female", "total"}
    header_row = first_val = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=200,
                                          values_only=True), start=1):
        hits = [i for i, v in enumerate(row)
                if v is not None and str(v).strip().lower() in genders]
        if len(hits) >= 3:
            header_row, first_val, n_val = ri, hits[0], len(hits)
            if hits != list(range(hits[0], hits[0] + n_val)):
                raise SystemExit(
                    f"{ws.title!r}: Male/Female/Total header cells are not "
                    f"contiguous (columns {hits}) — inspect the sheet manually."
                )
            break
    if header_row is None:
        raise SystemExit(
            f"{ws.title!r}: could not find a Male/Female/Total header row in the "
            f"first 10 rows. Run inspect_workbook.py to see the layout."
        )

    want = len(expected_groups) * 3
    if n_val != want:
        raise SystemExit(
            f"{ws.title!r}: found {n_val} value columns ({n_val // 3} groups × 3 "
            f"genders) but expected {want} ({len(expected_groups)} groups).\n"
            f"The published breakdown changed — update the group list in "
            f"clean_aishe.py to match before trusting the numbers.\n"
            f"Expected groups: {expected_groups}"
        )
    if first_val == 0:
        raise SystemExit(f"{ws.title!r}: value columns start at column A, leaving "
                         f"no label column — inspect the sheet manually.")
    label_col = first_val - 1

    # Verify the group labels sitting above the gender header (merged cells, so
    # only the top-left of each group carries a value).
    for ri in range(header_row - 1, max(0, header_row - 4), -1):
        row = next(ws.iter_rows(min_row=ri, max_row=ri, max_col=first_val + want,
                                values_only=True))
        labels = [str(v).strip() for i, v in enumerate(row)
                  if i >= first_val and v is not None and str(v).strip()]
        if len(labels) < 2:
            continue
        if len(labels) != len(expected_groups) or \
                [_key(x) for x in labels] != [_key(x) for x in expected_groups]:
            raise SystemExit(
                f"{ws.title!r}: group labels on row {ri} don't match the expected "
                f"order.\n  found:    {labels}\n  expected: {expected_groups}\n"
                f"Update the group list in clean_aishe.py (order matters — the "
                f"read is positional)."
            )
        break

    # First row after the header whose label column holds text.
    for ri in range(header_row + 1, min(header_row + 8, ws.max_row + 1)):
        v = ws.cell(row=ri, column=label_col + 1).value
        if v is not None and str(v).strip() and not isinstance(v, (int, float)):
            return label_col, first_val, ri
    raise SystemExit(f"{ws.title!r}: no data row found after header row {header_row}.")


# ─── Table 33: graduates by state × level ────────────────────────────────────
def state_level_rows(ws, year: str) -> list[dict]:
    label_col, first_val, data_row = _crosstab_geometry(ws, STATE_LEVEL_GROUPS)
    out = []
    grand: dict[str, int] = {}
    for row in ws.iter_rows(min_row=data_row, values_only=True):
        if len(row) <= label_col:
            continue
        state = _label_with_merge(row, label_col)
        if not state:
            continue
        state = canonical_state(state)
        is_rollup = _key(state) in ROLLUP_LABELS
        for li, group in enumerate(STATE_LEVEL_GROUPS):
            for gi, gender in enumerate(GENDERS):
                idx = first_val + li * 3 + gi
                val = row[idx] if idx < len(row) else None
                if is_rollup:
                    if group == GRAND_TOTAL:
                        grand[gender] = _num(val, f"{ws.title} AllIndia/{gender}")
                    continue
                if group == GRAND_TOTAL:
                    continue
                out.append(_row("state_level", year, "graduates", group,
                                state, SENTINEL, SENTINEL, SENTINEL, gender, val,
                                ctx=f"{ws.title} {state}/{group}/{gender}"))
    _check_grand_total(out, grand, year, ws.title)
    return out


def _check_grand_total(rows, grand, year: str, sheet: str) -> None:
    """The published Grand Total column must equal the sum of the levels.

    Free, and the strongest check available on this sheet: it catches a column
    misassignment that leaves every individual number looking plausible.
    """
    if not grand:
        print(f"  ⚠ {year} {sheet}: no All India row found — "
              f"grand-total check skipped")
        return
    for gender in GENDERS:
        got = sum(r["value"] for r in rows if r["gender"] == gender)
        want = grand.get(gender)
        if want is not None and got != want:
            raise SystemExit(
                f"{year} {sheet}: levels sum to {got:,} for gender={gender} but "
                f"the published Grand Total is {want:,} (off by {got - want:+,}).\n"
                f"A column was very likely misassigned — run "
                f"`python3 scripts/inspect_workbook.py --year {year}`."
            )


# ─── Table 34a: graduates by programme × social category (all levels) ────────
def programme_social_rows(ws, year: str) -> list[dict]:
    label_col, first_val, data_row = _crosstab_geometry(ws, SOCIAL_CATEGORIES)
    out = []
    for row in ws.iter_rows(min_row=data_row, values_only=True):
        if len(row) <= label_col:
            continue
        prog = _label_with_merge(row, label_col)
        if not prog or _key(prog) in ROLLUP_LABELS:
            continue
        for ci, cat in enumerate(SOCIAL_CATEGORIES):
            for gi, gender in enumerate(GENDERS):
                idx = first_val + ci * 3 + gi
                val = row[idx] if idx < len(row) else None
                out.append(_row("programme_social", year, "graduates",
                                SENTINEL, SENTINEL, SENTINEL, prog, cat, gender, val,
                                ctx=f"{ws.title} {prog}/{cat}/{gender}"))
    return out


# ─── Tables 12 / 35: UG by discipline ────────────────────────────────────────
def discipline_rows(ws, year: str, metric: str) -> list[dict]:
    """UG-by-discipline layout (shared by 12UGDisc enrolment and 35UGDisc
    graduates); shifts across years (S.No. column added in 2021-22)."""
    schema = None
    for ri in range(1, 6):
        cells = [str(c.value).strip() if c.value is not None else "" for c in ws[ri]]
        if cells and cells[0] == "Discipline":
            schema = "old"
            break
        if len(cells) >= 2 and cells[1] == "Discipline":
            schema = "new"
            break
    if schema is None:
        raise SystemExit(f"could not detect schema in sheet {ws.title!r}")
    if schema == "old":
        col_disc, col_subj, col_m, col_f, col_t = 0, 1, 2, 3, 4
    else:
        col_disc, col_subj, col_m, col_f, col_t = 1, 2, 3, 4, 5

    out = []
    published: dict[str, int] = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or len(r) <= col_t:
            continue
        subj, male, female, total = r[col_subj], r[col_m], r[col_f], r[col_t]
        disc_s = _label_with_merge(r, col_disc)
        if not disc_s:
            continue
        if disc_s.isdigit():
            continue
        is_total = (subj is None or str(subj).strip() == "") or disc_s.endswith("Total")
        if not is_total:
            continue
        clean = disc_s[:-len("Total")].strip() if disc_s.endswith("Total") else disc_s
        # AISHE spells the same discipline with a variable number of internal
        # spaces between editions ("Footwear  Design" in 2019-22 vs "Footwear
        # Design" in the PDF years), which silently splits it into two values and
        # breaks any trend or codemap join.
        clean = re.sub(r"\s+", " ", clean)
        if not clean or not isinstance(total, (int, float)):
            continue
        # The sheet's last row is the table's roll-up ("Grand Total", sometimes
        # "All India"). Stripping the "Total" suffix above turns it into a
        # discipline named "Grand", so it was being emitted as a real row — and
        # since it equals the sum of the disciplines, SUM(value) for 2019-20 and
        # 2020-21 returned exactly DOUBLE the true figure. Hold it back as the
        # validation anchor instead.
        if _key(clean) in ROLLUP_LABELS:
            ctx = f"{ws.title} roll-up"
            published = {"Male": _num(male, ctx), "Female": _num(female, ctx),
                         "Total": _num(total, ctx)}
            continue
        for gender, val in (("Male", male), ("Female", female), ("Total", total)):
            out.append(_row("ug_discipline", year, metric, "Under Graduate",
                            SENTINEL, clean, SENTINEL, SENTINEL, gender, val,
                            ctx=f"{ws.title} {clean}/{gender}"))
    _check_discipline_total(out, published, year, metric, ws.title)
    return out


def _check_discipline_total(rows, published, year: str, metric: str,
                            sheet: str) -> None:
    """The emitted disciplines must sum to the sheet's published roll-up.

    The check that would have caught the doubled 2019-20 / 2020-21 figures the
    moment they appeared.
    """
    if not published:
        print(f"  ⚠ {year} {sheet}: no roll-up row found — "
              f"discipline total unverified")
        return
    for gender in GENDERS:
        got = sum(r["value"] for r in rows if r["gender"] == gender)
        want = published[gender]
        if got != want:
            raise SystemExit(
                f"{year} {sheet} ({metric}): disciplines sum to {got:,} for "
                f"gender={gender} but the sheet's roll-up row says {want:,} "
                f"(off by {got - want:+,}).\nA roll-up or subject row is being "
                f"counted as a discipline — inspect the sheet before shipping."
            )


# ─── Driver ──────────────────────────────────────────────────────────────────
BUILDERS = {
    "state_level": lambda ws, year, metric: state_level_rows(ws, year),
    "programme_social": lambda ws, year, metric: programme_social_rows(ws, year),
    "ug_discipline": discipline_rows,
}


def build_rows(allow_missing_excel: bool = False) -> tuple[list[dict], list[str]]:
    """Parse every (year, sheet) pair declared in RAW_SHEETS.

    Returns (rows, missing_years). A missing workbook is only tolerated when
    `allow_missing_excel` is set; the caller decides what to do about it.
    """
    by_year: dict[str, list] = defaultdict(list)
    for rs in RAW_SHEETS:
        by_year[rs.year].append(rs)

    rows: list[dict] = []
    missing: list[str] = []
    for year, sheets in sorted(by_year.items()):
        if allow_missing_excel and not REPORTS[year].exists():
            missing.append(year)
            print(f"  {year} — workbook absent, skipped")
            continue
        wb = _wb(year)
        for rs in sheets:
            build = BUILDERS.get(rs.cut)
            if build is None:
                raise SystemExit(f"unknown cut {rs.cut!r} in RAW_SHEETS ({year}/{rs.sheet})")
            got = build(_sheet(wb, rs.sheet), year, rs.metric)
            print(f"  {year} {rs.sheet:<16} cut={rs.cut:<17} metric={rs.metric:<10} "
                  f"{len(got):>6,} rows")
            rows += got
    return rows, missing


def _validate(df: pd.DataFrame) -> None:
    for (cut, year), expected in sorted(ROW_BASELINE.items()):
        actual = len(df[(df.cut == cut) & (df.aishe_year == year)])
        if actual != expected:
            print(f"  ⚠ {cut} {year}: {actual:,} rows, baseline {expected:,} — "
                  f"the parse moved; investigate before shipping.")

    g = df[(df.metric == "graduates") & (df.gender == "Total")]
    years = sorted(set(g[g.cut == "state_level"].aishe_year) &
                   set(g[g.cut == "ug_discipline"].aishe_year))
    if not years:
        print("  (no year carries both the state_level and ug_discipline cuts — "
              "cross-cut reconciliation skipped)")
    for year in years:
        ug_state = g[(g.cut == "state_level") & (g.aishe_year == year)
                     & (g.level == "Under Graduate")].value.sum()
        ug_disc = g[(g.cut == "ug_discipline") & (g.aishe_year == year)].value.sum()
        anchor = UG_GRADUATES_ANCHOR.get(year)
        ok = ug_state == ug_disc and (anchor is None or ug_state == anchor)
        note = f"  anchor={anchor:,}" if anchor else "  (no published anchor)"
        print(f"  {year} UG graduates: state-cut={ug_state:,}  "
              f"discipline-cut={ug_disc:,}{note}  {'OK' if ok else 'CHECK'}")


def _check_programme_vs_state(df: pd.DataFrame) -> None:
    """Table 34's population is Table 33's Grand Total — so they must agree.

    An external anchor for the programme cut that costs nothing: the programme
    table and the state×level table count the same graduates, sliced differently.
    It is what confirmed the 2018-19 Table 34 read after the sparse-row fix, and
    it is the check to lean on for any edition that prints no Grand Total row of
    its own (2015-16, 2016-17).
    """
    g = df[df.social_category.isin(["All Categories"]) | (df.cut == "state_level")]
    for year in sorted(set(df[df.cut == "programme_social"].aishe_year)):
        if year not in set(df[df.cut == "state_level"].aishe_year):
            continue
        for gender in GENDERS:
            a = g[(g.cut == "state_level") & (g.aishe_year == year)
                  & (g.gender == gender)].value.sum()
            b = g[(g.cut == "programme_social") & (g.aishe_year == year)
                  & (g.gender == gender)
                  & (g.social_category == "All Categories")].value.sum()
            if a != b:
                raise SystemExit(
                    f"{year}: programme cut sums to {b:,} for gender={gender} but "
                    f"the state×level cut sums to {a:,} (off by {b - a:+,}).\n"
                    f"These count the same graduates — one of the two reads is "
                    f"wrong."
                )
        print(f"  {year} programme vs state×level: agree on all genders  OK")


def _check_state_labels(df: pd.DataFrame) -> None:
    """Every year of a state cut must carry the same states, one row-count each.

    The reconciliation checks compare sums, so a mangled state *name* passes them
    silently: a label split in two still totals correctly. That is exactly how the
    2018-19 Table 33 "Chhattisgarh Dadra and Nagar" / "Haveli" split reached
    production. This checks the shape of the dimension instead of its sums.
    """
    for cut in ("state_level", "state_social"):
        sub = df[df.cut == cut]
        if sub.empty:
            continue
        for year, g in sub.groupby("aishe_year"):
            counts = g.groupby("state").size()
            if counts.nunique() != 1:
                odd = counts[counts != counts.mode()[0]].to_dict()
                raise SystemExit(
                    f"{cut} {year}: states do not all have the same number of "
                    f"rows — {odd}.\nA state label was split or merged by the row "
                    f"reader. The totals can still reconcile, so this is the only "
                    f"check that catches it."
                )


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--allow-missing-excel", action="store_true",
                    help="Build from whatever is present instead of failing. "
                         "Writes higher_ed.partial.parquet, which is never "
                         "uploaded or loaded.")
    args = ap.parse_args()

    print(f"AISHE parse — Excel: {len(RAW_SHEETS)} sheets across "
          f"{len({rs.year for rs in RAW_SHEETS})} years; "
          f"PDF: {len(PDF_TABLES)} tables across "
          f"{len({t.year for t in PDF_TABLES})} years")

    excel_rows, missing = build_rows(args.allow_missing_excel)
    pdf_rows = parse_report_pdf.build_rows()

    df = pd.DataFrame(excel_rows + pdf_rows, columns=COLUMNS)
    df["value"] = df["value"].astype("Int64")

    CLEAN.mkdir(parents=True, exist_ok=True)
    # A build that skipped a registered Excel year is not the whole table, and
    # load_bq.py truncates — so it must not be able to reach the loader. Routing
    # it to a different filename makes that a property of the pipeline rather
    # than something the next person has to remember.
    partial = bool(missing)
    out = (CLEAN / "higher_ed.partial.parquet") if partial else TABLES[0].local_path
    df.to_parquet(out, index=False, engine="pyarrow")

    if partial:
        print(f"\n⚠ PARTIAL BUILD — missing Excel year(s): {', '.join(missing)}")
        print(f"  Wrote {out.name}, which upload_to_gcs.py and load_bq.py ignore.")
        print(f"  Restore the workbooks first:  python3 scripts/fetch.py --from-gcs")

    print(f"\nAISHE → {out.name}: {len(df):,} rows")
    for cut in sorted(df.cut.unique()):
        sub = df[df.cut == cut]
        if sub.empty:
            continue
        by_metric = ", ".join(f"{m}={(sub.metric == m).sum():,}"
                              for m in ("graduates", "enrolment") if (sub.metric == m).any())
        yrs = ",".join(sorted(sub.aishe_year.unique()))
        bases = "/".join(sorted(sub.basis.unique()))
        print(f"  cut={cut:<17} {len(sub):>6,} rows  ({by_metric})  "
              f"basis={bases:<16} years={yrs}")

    _validate(df)
    _check_programme_vs_state(df)
    _check_state_labels(df)
    print("✓ done.")


if __name__ == "__main__":
    main()
