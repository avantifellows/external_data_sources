#!/usr/bin/env python3
"""
Inspect the AISHE Final Report workbooks in raw/ — sheet inventory plus header
geometry for the sheets clean_aishe.py parses.

Run this FIRST when adding a new AISHE year, before trusting clean_aishe.py.

Why: the parser is only partly self-describing. `_discipline_series` (Tables 12
and 35) detects its header by locating the "Discipline" cell, so it tolerates a
column shift. Tables 33 and 34a do NOT — they read fixed offsets
(`row[2 + i * 3 + gender]`) from a fixed start row, against fixed LEVELS and
SOCIAL_CATEGORIES lists. And `_row()` coerces a non-numeric cell to 0. So a
workbook that renumbers a table, inserts an S.No. column, or adds a social
category yields plausible zeros and wrong totals instead of an error.

This script surfaces the three things that decide whether the parser can be
reused as-is for a new year:
  1. Does each sheet clean_aishe.py wants still exist (and under what name)?
  2. Where does the data actually start, and what is in the leading columns?
  3. How many value columns follow the label columns — i.e. does the
     level / social-category count still match the hardcoded lists?

Usage:
  python3 scripts/inspect_workbook.py                       # every year present in raw/
  python3 scripts/inspect_workbook.py --year 2022-23
  python3 scripts/inspect_workbook.py --year 2022-23 --rows 8 --cols 14
  python3 scripts/inspect_workbook.py --all-sheets          # list every sheet, not just wanted
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from sources import RAW_SHEETS, REPORTS

# The sheets clean_aishe.py parses, taken from the RAW_SHEETS registry so this
# script can't drift from the pipeline.
WANTED: list[str] = sorted({rs.sheet for rs in RAW_SHEETS})


def _norm(name: str) -> str:
    """clean_aishe._sheet's matching rule: space-stripped, lowercased."""
    return name.replace(" ", "").lower()


def _alpha(name: str) -> str:
    return re.sub(r"[^a-z]", "", name.lower())


def _digits(name: str) -> str:
    return re.sub(r"[^0-9]", "", name)


def _candidates(wanted: str, sheetnames: list[str]) -> list[str]:
    """Sheets a missing `wanted` may have been renamed to.

    AISHE renumbers tables between editions, so the non-numeric part is the
    stable half: "35UGDisc" -> stem "ugdisc" matches a "36UGDisc". Short stems
    ("34a" -> "a") are weak on their own, so also offer sheets carrying the same
    table number. Both are suggestions for a human to confirm, so a loose match
    is better than none.
    """
    stem = _alpha(wanted)
    out = [s for s in sheetnames if stem and _alpha(s) == stem]
    if len(stem) < 3:
        out += [s for s in sheetnames if _digits(s) == _digits(wanted) and s not in out]
    return out


def _preview(ws, n_rows: int, n_cols: int) -> None:
    """Print the top-left corner of the sheet, one line per row."""
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=n_rows,
                                          max_col=n_cols, values_only=True), start=1):
        cells = []
        for v in row:
            if v is None:
                cells.append("·")
            else:
                s = str(v).strip().replace("\n", " ")
                cells.append(s[:14] if len(s) <= 14 else s[:13] + "…")
        print(f"      r{ri:<2} | " + " | ".join(cells))


def _geometry(ws, label_col: int, max_scan: int = 12, max_col: int = 120):
    """Locate the first data row for `label_col` and measure its value block.

    Returns (data_row, first_value_col, n_value_cols) or None.

    A data row is one where `label_col` holds non-numeric text and at least one
    later cell is numeric. The value block does not necessarily start adjacent
    to the label — the discipline sheets carry a "Subject" column in between —
    so the first numeric column is reported rather than assumed.

    clean_aishe.py hardcodes min_row=5 for Tables 33/34a and min_row=4 for the
    discipline sheets; this reports what the workbook actually has.
    """
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan,
                                          max_col=max_col, values_only=True), start=1):
        if len(row) <= label_col:
            continue
        label = row[label_col]
        if label is None or not str(label).strip() or isinstance(label, (int, float)):
            continue
        nums = [i for i, v in enumerate(row) if i > label_col and isinstance(v, (int, float))]
        if nums:
            return ri, nums[0], len(nums)
    return None


def inspect(year: str, path: Path, n_rows: int, n_cols: int, all_sheets: bool) -> bool:
    print(f"\n{'=' * 78}\n{year}  —  {path.name}  ({path.stat().st_size / 1e6:.2f} MB)\n{'=' * 78}")
    wb = openpyxl.load_workbook(path, data_only=True)
    names = wb.sheetnames
    print(f"  {len(names)} sheets")
    if all_sheets:
        for s in names:
            print(f"    · {s}")

    lookup = {_norm(s): s for s in names}
    expected = {rs.sheet for rs in RAW_SHEETS if rs.year == year}
    if not expected:
        print(f"  (year not yet in RAW_SHEETS — checking all {len(WANTED)} known sheets)")

    all_found = True
    for wanted in WANTED:
        tag = "" if not expected or wanted in expected else "  [not registered for this year]"
        actual = lookup.get(_norm(wanted))
        if actual is None:
            all_found = False
            cands = _candidates(wanted, names)
            hint = f"  candidates: {cands}" if cands else "  no obvious rename candidate"
            print(f"\n  ✗ {wanted:<16} MISSING{tag}{hint}")
            continue

        print(f"\n  ✓ {wanted:<16} → sheet {actual!r}{tag}")
        ws = wb[actual]
        print(f"      dims {ws.max_row} rows × {ws.max_column} cols")
        # Tables 33/34a label in col B (index 1); discipline sheets may be A or B.
        for label_col in (1, 0):
            geo = _geometry(ws, label_col)
            if geo is not None:
                dr, first_val, n_val = geo
                groups = (f"{n_val // 3} groups × 3 genders" if n_val % 3 == 0
                          else "NOT a multiple of 3 — genders may not be triples")
                print(f"      data starts r{dr}; label col {'AB'[label_col]} (index {label_col}); "
                      f"{n_val} value cols from index {first_val} → {groups}")
                break
        else:
            print("      could not locate a data row in the first 12 rows — inspect manually")
        _preview(ws, n_rows, n_cols)

    return all_found


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--year", default=None, help="Inspect only this year (e.g. 2022-23)")
    ap.add_argument("--rows", type=int, default=6, help="Preview rows per sheet (default 6)")
    ap.add_argument("--cols", type=int, default=12, help="Preview cols per sheet (default 12)")
    ap.add_argument("--all-sheets", action="store_true",
                    help="List every sheet name, not just the ones the parser wants")
    args = ap.parse_args()

    years = [args.year] if args.year else list(REPORTS)
    for year in years:
        if year not in REPORTS:
            raise SystemExit(f"unknown year {year!r}; known: {list(REPORTS)}\n"
                             f"Add it to REPORTS / REPORT_URLS in sources.py first.")

    present = [(y, REPORTS[y]) for y in years if REPORTS[y].exists()]
    missing = [y for y in years if not REPORTS[y].exists()]
    if missing:
        print(f"not in raw/ (run fetch.py): {', '.join(missing)}")
    if not present:
        raise SystemExit("nothing to inspect — no workbooks in raw/.")

    ok = all(inspect(y, p, args.rows, args.cols, args.all_sheets) for y, p in present)
    print(f"\n{'✓ all wanted sheets found.' if ok else '✗ some sheets missing — see above.'}")


if __name__ == "__main__":
    main()
